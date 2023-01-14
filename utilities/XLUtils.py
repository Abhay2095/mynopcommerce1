import openpyxl


def getRowCount(file, sheetname):
    worksheet = openpyxl.load_workbook(file)
    sheet = worksheet[sheetname]
    return sheet.max_row


def getColumnCount(file, sheetname):
    worksheet = openpyxl.load_workbook(file)
    sheet = worksheet[sheetname]
    return sheet.max_column


def readData(file, sheetname, rownum, columnno):
    worksheet = openpyxl.load_workbook(file)
    sheet = worksheet[sheetname]
    return sheet.cell(row=rownum, column=columnno).value


def writeData(file, sheetname, rownum, columnno):
    worksheet = openpyxl.load_workbook(file)
    sheet = worksheet[sheetname]
    return sheet.cell(row=rownum, column=columnno).value
    worksheet.save(file)




